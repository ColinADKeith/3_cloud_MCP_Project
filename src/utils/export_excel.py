import os
import sys
import json
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

# Path configuration for standard workspace routing
root_workspace = os.path.abspath(os.path.join(os.path.dirname(__file__), "../.."))
if root_workspace not in sys.path:
    sys.path.insert(0, root_workspace)

def generate_production_excel():
    json_path = "data/qualified_jobs.json"
    output_xlsx = "data/Qualified_Job_Pipeline.xlsx"
    
    if not os.path.exists(json_path):
        print(f"❌ Error: Missing source file '{json_path}'. Run your screener agent first!")
        return
        
    with open(json_path, "r", encoding="utf-8") as f:
        jobs = json.load(f)
        
    wb = openpyxl.Workbook()
    
    # TAB 1: DASHBOARD COMPILATION
    ws_dash = wb.active
    ws_dash.title = "Summary Dashboard"
    ws_dash.views.sheetView[0].showGridLines = True
    
    PRIMARY_COLOR = "1F4E78"
    ACCENT_COLOR = "D9E1F2"
    fill_header = PatternFill(fill_type="solid", start_color=PRIMARY_COLOR, end_color=PRIMARY_COLOR)
    fill_accent = PatternFill(fill_type="solid", start_color=ACCENT_COLOR, end_color=ACCENT_COLOR)
    fill_zebra = PatternFill(fill_type="solid", start_color="F9FBFD", end_color="F9FBFD")
    thin_border = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
                         top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))
    
    ws_dash.cell(row=2, column=2, value="🕵️ AGENT CONTROL CENTER: PIPELINE ANALYTICS").font = Font(size=16, bold=True, color=PRIMARY_COLOR)
    ws_dash.cell(row=3, column=2, value="Real-Time Sourced & Mathematically Filtered Targets").font = Font(size=10, italic=True, color="595959")
    
    kpi_labels = ["TOTAL PIPELINE JOBS", "AVERAGE MATCH PROXIMITY"]
    kpi_vals = [len(jobs), "=AVERAGE('Qualified Jobs List'!F2:F5000)"]
    kpi_cols = [("B", "C"), ("E", "F")]
    
    for label, val, (sc, ec) in zip(kpi_labels, kpi_vals, kpi_cols):
        cell_lbl = ws_dash.cell(row=5, column=openpyxl.utils.column_index_from_string(sc), value=label)
        cell_lbl.font = Font(size=9, bold=True, color="FFFFFF")
        cell_lbl.fill = fill_header
        cell_lbl.alignment = Alignment(horizontal="center")
        ws_dash.merge_cells(f"{sc}5:{ec}5")
        
        cell_val = ws_dash.cell(row=6, column=openpyxl.utils.column_index_from_string(sc), value=val)
        cell_val.font = Font(size=14, bold=True, color=PRIMARY_COLOR)
        cell_val.fill = fill_accent
        cell_val.alignment = Alignment(horizontal="center")
        if "%" not in str(val) and "=" in str(val):
            cell_val.number_format = "0.0000"
        ws_dash.merge_cells(f"{sc}6:{ec}6")
        
    counts = {}
    for j in jobs:
        src = j.get("source", "Unknown Board")
        counts[src] = counts.get(src, 0) + 1
        
    ws_dash.cell(row=9, column=2, value="CHANNEL SOURCE METRICS").font = Font(size=11, bold=True, color=PRIMARY_COLOR)
    headers = ["Job Acquisition Channel", "Aggregated Openings", "Market Share %"]
    for c_idx, h in enumerate(headers, 2):
        c = ws_dash.cell(row=10, column=c_idx, value=h)
        c.font = Font(color="FFFFFF", bold=True); c.fill = fill_header; c.alignment = Alignment(horizontal="center")
        
    for r_idx, (src, count) in enumerate(counts.items(), 11):
        ws_dash.cell(row=r_idx, column=2, value=src).alignment = Alignment(horizontal="left")
        ws_dash.cell(row=r_idx, column=3, value=count).number_format = "#,##0"
        ws_dash.cell(row=r_idx, column=4, value=f"=C{r_idx}/C{11+len(counts)}").number_format = "0.0%"
        
        for c in range(2, 5):
            cell = ws_dash.cell(row=r_idx, column=c)
            cell.font = Font(size=10); cell.border = thin_border
            if r_idx % 2 == 0: cell.fill = fill_zebra
            
    tot_row = 11 + len(counts)
    ws_dash.cell(row=tot_row, column=2, value="Total Target Spectrum").font = Font(bold=True)
    ws_dash.cell(row=tot_row, column=3, value=f"=SUM(C11:C{tot_row-1})").font = Font(bold=True)
    ws_dash.cell(row=tot_row, column=3).number_format = "#,##0"
    ws_dash.cell(row=tot_row, column=4, value=f"=SUM(D11:D{tot_row-1})").font = Font(bold=True)
    ws_dash.cell(row=tot_row, column=4).number_format = "0.0%"
    
    # TAB 2: DETAILED RECORD EXTRACTION
    ws_jobs = wb.create_sheet(title="Qualified Jobs List")
    ws_jobs.views.sheetView[0].showGridLines = True
    
    # 💥 ADDED: "Date Added" Column Header
    data_headers = ["Job ID", "Job Title", "Company", "Location", "Source", "Semantic Distance", "Portal Link URL", "Date Added"]
    for col_idx, h in enumerate(data_headers, 1):
        cell = ws_jobs.cell(row=1, column=col_idx, value=h)
        cell.font = Font(color="FFFFFF", bold=True); cell.fill = fill_header; cell.alignment = Alignment(horizontal="center")
        
    for r_idx, j in enumerate(jobs, 2):
        ws_jobs.cell(row=r_idx, column=1, value=j.get("job_id")).alignment = Alignment(horizontal="center")
        ws_jobs.cell(row=r_idx, column=2, value=j.get("title"))
        ws_jobs.cell(row=r_idx, column=3, value=j.get("company"))
        ws_jobs.cell(row=r_idx, column=4, value=j.get("location"))
        ws_jobs.cell(row=r_idx, column=5, value=j.get("source"))
        
        dist_cell = ws_jobs.cell(row=r_idx, column=6, value=j.get("semantic_distance", 2.0))
        dist_cell.number_format = "0.0000"
        dist_cell.alignment = Alignment(horizontal="right")
        
        url_cell = ws_jobs.cell(row=r_idx, column=7, value=j.get("url"))
        url_cell.font = Font(color="0563C1", underline="single")
        
        # 💥 ADDED: Populate "Date Added" Cell Values
        date_cell = ws_jobs.cell(row=r_idx, column=8, value=j.get("date_added", datetime.date.today().strftime("%Y-%m-%d")))
        date_cell.alignment = Alignment(horizontal="center")
        
        for c in range(1, 9): # Expanded loop boundary to 9 to encompass the date column
            cell = ws_jobs.cell(row=r_idx, column=c)
            cell.font = Font(size=10); cell.border = thin_border
            if r_idx % 2 == 0: cell.fill = fill_zebra
            
    rule = ColorScaleRule(start_type='num', start_value=0.28, start_color='E2EFDA',
                          end_type='num', end_value=0.55, end_color='FFF2CC')
    ws_jobs.conditional_formatting.add(f"F2:F{len(jobs)+1}", rule)
    
    for col in ws_jobs.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_jobs.column_dimensions[col_letter].width = min(max(max_len + 3, 11), 40)
        
    ws_jobs.auto_filter.ref = f"A1:H{len(jobs)+1}"
    ws_jobs.freeze_panes = "A2"
    
    wb.save(output_xlsx)
    print(f"🏁 Clean File Export Complete: Created '{output_xlsx}' containing {len(jobs)} active entries.")

if __name__ == "__main__":
    generate_production_excel()